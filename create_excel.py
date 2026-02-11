#!/usr/bin/env python3
import csv
import openpyxl
from openpyxl.styles import Font, Alignment
import os

def csv_to_excel(csv_path, excel_path):
    """将CSV文件转换为Excel文件"""
    try:
        # 读取CSV文件
        with open(csv_path, 'r', encoding='utf-8-sig') as csvfile:
            csv_reader = csv.reader(csvfile)
            rows = list(csv_reader)
        
        if not rows:
            print("CSV文件为空")
            return False
        
        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "学生数据汇总"
        
        # 设置标题样式
        header_font = Font(bold=True, size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # 写入数据
        for row_idx, row in enumerate(rows, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # 标题行
                    cell.font = header_font
                    cell.alignment = header_alignment
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存Excel文件
        wb.save(excel_path)
        print(f"Excel文件已创建: {excel_path}")
        return True
        
    except Exception as e:
        print(f"创建Excel文件时出错: {e}")
        return False

def create_comprehensive_excel(csv_path, excel_path):
    """创建更完善的Excel文件"""
    try:
        # 读取CSV数据
        with open(csv_path, 'r', encoding='utf-8-sig') as csvfile:
            csv_reader = csv.DictReader(csvfile)
            students = list(csv_reader)
        
        if not students:
            print("没有学生数据")
            return False
        
        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        
        # 1. 主数据表
        ws_main = wb.active
        ws_main.title = "学生汇总"
        
        # 写入标题
        headers = ['学校', '姓名', '年龄', '邮箱', '高中', '序号', '出生日期']
        for col_idx, header in enumerate(headers, 1):
            cell = ws_main.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center')
        
        # 写入学生数据
        for row_idx, student in enumerate(students, 2):
            ws_main.cell(row=row_idx, column=1, value=student.get('学校', ''))
            ws_main.cell(row=row_idx, column=2, value=student.get('姓名', ''))
            ws_main.cell(row=row_idx, column=3, value=student.get('年龄', ''))
            ws_main.cell(row=row_idx, column=4, value=student.get('邮箱', ''))
            ws_main.cell(row=row_idx, column=5, value=student.get('高中', ''))
            ws_main.cell(row=row_idx, column=6, value=student.get('序号', ''))
            ws_main.cell(row=row_idx, column=7, value=student.get('出生日期', ''))
        
        # 2. 学校统计表
        ws_stats = wb.create_sheet(title="学校统计")
        
        # 按学校统计
        school_counts = {}
        for student in students:
            school = student.get('学校', '未知')
            school_counts[school] = school_counts.get(school, 0) + 1
        
        # 写入统计标题
        ws_stats.cell(row=1, column=1, value="学校").font = Font(bold=True)
        ws_stats.cell(row=1, column=2, value="学生人数").font = Font(bold=True)
        ws_stats.cell(row=1, column=3, value="占比").font = Font(bold=True)
        
        # 写入统计数据
        total_students = len(students)
        for row_idx, (school, count) in enumerate(sorted(school_counts.items()), 2):
            percentage = (count / total_students) * 100
            
            ws_stats.cell(row=row_idx, column=1, value=school)
            ws_stats.cell(row=row_idx, column=2, value=count)
            ws_stats.cell(row=row_idx, column=3, value=f"{percentage:.1f}%")
        
        # 调整列宽
        for ws in [ws_main, ws_stats]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存文件
        wb.save(excel_path)
        
        print(f"Excel文件已创建: {excel_path}")
        print(f"总学生数: {len(students)}")
        print("包含工作表:")
        print("  1. '学生汇总' - 所有学生详细信息")
        print("  2. '学校统计' - 按学校分布统计")
        
        return True
        
    except Exception as e:
        print(f"创建Excel文件时出错: {e}")
        return False

def main():
    csv_path = "/Users/yangyan/.openclaw/workspace/helenoa_students.csv"
    excel_path = "/Users/yangyan/.openclaw/workspace/HelenOA_学生数据汇总.xlsx"
    
    if not os.path.exists(csv_path):
        print(f"CSV文件不存在: {csv_path}")
        return
    
    print("正在创建Excel文件...")
    if create_comprehensive_excel(csv_path, excel_path):
        print(f"\n文件已成功创建:")
        print(f"  {excel_path}")
        
        # 显示文件大小
        if os.path.exists(excel_path):
            size_bytes = os.path.getsize(excel_path)
            size_kb = size_bytes / 1024
            print(f"  文件大小: {size_kb:.1f} KB")
    else:
        print("创建Excel文件失败")

if __name__ == "__main__":
    main()